from fastapi import FastAPI, Body
from pydantic import BaseModel, Field
from typing import List
from openpyxl import Workbook, load_workbook
from datetime import datetime
from fastapi.middleware.cors import CORSMiddleware
import os

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"], 
    allow_headers=["*"], 
)

FILE_NAME = "reacoes.xlsx"

class Votes(BaseModel):
    Amor: int
    Tomate: int
    Engraçado: int
    voce_0: int = Field(alias="Você é 0")
    voce_10: int = Field(alias="Você é 10")
    coracao_partido: int = Field(alias="Coração Partido")

class Person(BaseModel):
    id: int
    name: str
    votes: Votes

def create_file_if_not_exists():
    if not os.path.exists(FILE_NAME):
        wb = Workbook()
        ws = wb.active
        ws.title = "Reacoes"
        ws.append(["Pessoa", "Reacao", "DataHora"])
        wb.save(FILE_NAME)

def save_to_excel(person_name: str, reaction: str):
    create_file_if_not_exists()

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    ws.append([
        person_name,
        reaction,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ])

    wb.save(FILE_NAME)

def read_and_group_data():
    create_file_if_not_exists()

    wb = load_workbook(FILE_NAME)
    ws = wb.active

    result = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        person, reaction, date = row

        if person not in result:
            result[person] = {
                "votes": {"Amor": 0, "Tomate": 0, "Engraçado": 0, "Você é 0": 0, "Você é 10": 0, "Coração Partido": 0},
                "last_update": date
            }

        if reaction in result[person]["votes"]:
            result[person]["votes"][reaction] += 1

        result[person]["last_update"] = date

    people_list = []
    for i, (name, data) in enumerate(result.items(), start=1):
        people_list.append({
            "id": i,
            "name": name,
            "votes": data["votes"],
            "last_update": data["last_update"]
        })

    people_list.sort(key=lambda x: x["last_update"], reverse=True)

    return people_list

@app.get("/people", response_model=List[Person])
async def get_people():
    return read_and_group_data()

@app.post("/send-reaction")
async def send_reaction(data: dict = Body(...)):
    person_name = data.get("person_name")
    reaction = data.get("reaction")

    valid_reactions = ["Amor", "Tomate", "Engraçado", "Você é 0", "Você é 10", "Coração Partido"]

    if not person_name:
        return {"error": "Nome obrigatório"}

    if reaction not in valid_reactions:
        return {"error": "Reação inválida"}

    save_to_excel(person_name, reaction)

    updated_data = read_and_group_data()

    return {
        "message": "Reação salva",
        "data": updated_data
    }